from openpyxl import Workbook
import openpyxl

wb = Workbook()
sheet = wb.active
sheet['A1'] = "Hello"  
sheet['B1'] = "World!"  
sheet['A3'] = 41.80  
sheet['A4'] = 10

wb.save("example.xlsx")  

wb = openpyxl.load_workbook("example.xlsx") 
sheet = wb.active
temp1 = sheet['A1']
temp2 = sheet['B1']
temp3 = sheet.cell(row = 3, column = 1)
temp4 = sheet.cell(row = 4, column = 1)
print(temp1.value, temp2.value, temp3.value, temp4.value)
